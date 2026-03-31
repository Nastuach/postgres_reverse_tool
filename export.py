import os
from datetime import datetime
from tkinter import filedialog, messagebox
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ExportManager:
    """Менеджер экспорта данных"""

    def __init__(self, app):
        self.app = app

    def export_to_pdf(self):
        """Экспорт в PDF"""
        if not self.app.current_data:
            messagebox.showwarning("Нет данных", "Сначала выполните анализ базы данных!")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"db_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not filename:
            return

        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                        fontSize=16, textColor=colors.HexColor('#7c3aed'))
            story.append(Paragraph("Отчет о структуре базы данных PostgreSQL", title_style))
            story.append(Spacer(1, 0.2*inch))

            stats_data = [['Показатель', 'Значение']]
            for name, value in self.app.current_data['stats'].items():
                stats_data.append([name, str(value)])

            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 0.3*inch))

            doc.build(story)
            messagebox.showinfo("Успех", f"PDF-отчет сохранен!\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF: {e}")

    def export_to_excel(self):
        """Экспорт в Excel"""
        if not self.app.current_data:
            messagebox.showwarning("Нет данных", "Сначала выполните анализ базы данных!")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"db_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "Статистика"
            ws['A1'] = "Статистика базы данных"
            ws['A1'].font = Font(size=14, bold=True)

            row = 3
            for name, value in self.app.current_data['stats'].items():
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=value)
                row += 1

            ws_tables = wb.create_sheet("Таблицы")
            ws_tables['A1'] = "Таблицы базы данных"
            ws_tables['A1'].font = Font(size=14, bold=True)
            ws_tables['A3'] = "Таблица"
            ws_tables['B3'] = "Колонок"
            ws_tables['C3'] = "Строк"
            ws_tables['D3'] = "Описание"

            row = 4
            for table_name, table in self.app.current_data['tables'].items():
                ws_tables.cell(row=row, column=1, value=table_name)
                ws_tables.cell(row=row, column=2, value=len(table.columns))
                ws_tables.cell(row=row, column=3, value=table.row_count or 0)
                ws_tables.cell(row=row, column=4, value=table.description or '')
                row += 1

            ws_columns = wb.create_sheet("Колонки")
            ws_columns['A1'] = "Колонки таблиц"
            ws_columns['A1'].font = Font(size=14, bold=True)
            headers = ['Таблица', 'Колонка', 'Тип', 'Nullable', 'Default', 'Вычисляемое', 'Описание']
            for col, header in enumerate(headers, 1):
                ws_columns.cell(row=3, column=col, value=header).font = Font(bold=True)

            row = 4
            for table_name, table in self.app.current_data['tables'].items():
                for col in table.columns:
                    ws_columns.cell(row=row, column=1, value=table_name)
                    ws_columns.cell(row=row, column=2, value=col.name)
                    ws_columns.cell(row=row, column=3, value=col.data_type)
                    ws_columns.cell(row=row, column=4, value='Да' if col.is_nullable else 'Нет')
                    ws_columns.cell(row=row, column=5, value=col.default_value or '')
                    ws_columns.cell(row=row, column=6, value='Да' if col.is_computed else 'Нет')
                    ws_columns.cell(row=row, column=7, value=col.description or '')
                    row += 1

            ws_fk = wb.create_sheet("Связи")
            ws_fk['A1'] = "Внешние ключи"
            ws_fk['A1'].font = Font(size=14, bold=True)
            ws_fk['A3'] = "Связь"
            ws_fk['B3'] = "Таблица"
            ws_fk['C3'] = "Колонки"
            ws_fk['D3'] = "Ссылается на"

            row = 4
            for table_name, table in self.app.current_data['tables'].items():
                for fk in table.get_foreign_keys():
                    ws_fk.cell(row=row, column=1, value=fk.name)
                    ws_fk.cell(row=row, column=2, value=table_name)
                    ws_fk.cell(row=row, column=3, value=', '.join(fk.columns))
                    ws_fk.cell(row=row, column=4, value=f"{fk.referenced_table}({', '.join(fk.referenced_columns)})" if fk.referenced_table else '')
                    row += 1

            wb.save(filename)
            messagebox.showinfo("Успех", f"Excel-отчет сохранен!\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать Excel: {e}")